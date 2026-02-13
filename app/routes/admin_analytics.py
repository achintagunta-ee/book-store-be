from fastapi import APIRouter, Depends
from sqlmodel import Session, func, select
from app.database import get_session
from app.models.book import Book
from app.models.category import Category
from app.models.order import Order
from app.models.order_item import OrderItem
from app.models.user import User
from fastapi.responses import StreamingResponse
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference




router = APIRouter()
@router.get("/overview")
def analytics_overview(session: Session = Depends(get_session)):
    total_revenue = session.exec(
        select(func.sum(Order.total))
        .where(Order.status == "paid")
    ).one()

    total_orders = session.exec(
        select(func.count(Order.id))
        .where(Order.status == "paid")
    ).one()

    avg_order = total_revenue / total_orders if total_orders else 0

    return {
        "revenue": total_revenue or 0,
        "orders": total_orders or 0,
        "avg_order_value": avg_order
    }

@router.get("/revenue-chart")
def revenue_chart(session: Session = Depends(get_session)):
    data = session.exec(
        select(
            func.date(Order.created_at),
            func.sum(Order.total)
        )
        .where(Order.status == "paid")
        .group_by(func.date(Order.created_at))
        .order_by(func.date(Order.created_at))
    ).all()

    return [
        {"date": str(d), "revenue": total}
        for d, total in data
    ]

@router.get("/top-books")
def top_books(session: Session = Depends(get_session)):
    data = session.exec(
        select(
            OrderItem.book_title,
            func.sum(OrderItem.quantity).label("sold")
        )
        .group_by(OrderItem.book_title)
        .order_by(func.sum(OrderItem.quantity).desc())
        .limit(5)
    ).all()

    return [{"title": t, "sold": s} for t, s in data]

@router.get("/top-customers")
def top_customers(session: Session = Depends(get_session)):
    data = session.exec(
        select(
            User.email,
            func.sum(Order.total).label("spent"),
            func.count(Order.id).label("orders")
        )
        .join(Order, Order.user_id == User.id)
        .where(Order.status == "paid")
        .group_by(User.email)
        .order_by(func.sum(Order.total).desc())
        .limit(5)
    ).all()

    return [
        {"email": email, "spent": spent, "orders": orders}
        for email, spent, orders in data
    ]

@router.get("/category-sales")
def category_sales(session: Session = Depends(get_session)):
    data = session.exec(
        select(
            Category.name,
            func.sum(OrderItem.quantity)
        )
        .join(Book, Book.category_id == Category.id)
        .join(OrderItem, OrderItem.book_id == Book.id)
        .group_by(Category.name)
    ).all()

    return [{"category": c, "sold": s} for c, s in data]



@router.get("/export")
def export_excel(session: Session = Depends(get_session)):

    wb = Workbook()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center")

    currency = "₹#,##0.00"

    # =========================
    # Sheet 1 — Overview
    # =========================
    ws = wb.active
    ws.title = "Overview"

    ws.append(["Metric", "Value"])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    total_revenue = session.exec(
        select(func.sum(Order.total)).where(Order.status == "paid")
    ).one() or 0

    total_orders = session.exec(
        select(func.count(Order.id)).where(Order.status == "paid")
    ).one() or 0

    avg = total_revenue / total_orders if total_orders else 0

    rows = [
        ["Total Revenue", total_revenue],
        ["Total Orders", total_orders],
        ["Average Order Value", avg],
    ]

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2):
        row[1].number_format = currency
        for cell in row:
            cell.border = thin
            cell.alignment = center

    # =========================
    # Sheet 2 — Revenue by Day
    # =========================
    ws2 = wb.create_sheet("Revenue")

    ws2.append(["Date", "Revenue"])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    data = session.exec(
        select(func.date(Order.created_at), func.sum(Order.total))
        .where(Order.status == "paid")
        .group_by(func.date(Order.created_at))
        .order_by(func.date(Order.created_at))
    ).all()

    for d, total in data:
        ws2.append([d, total])

    for row in ws2.iter_rows(min_row=2):
        row[1].number_format = currency
        for cell in row:
            cell.border = thin
            cell.alignment = center

    # Chart
    chart = BarChart()
    chart.title = "Revenue Trend"

    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(data)+1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(data)+1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    ws2.add_chart(chart, "D3")

    # =========================
    # Sheet 3 — Top Books
    # =========================
    ws3 = wb.create_sheet("Top Books")
    ws3.append(["Title", "Units Sold"])

    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    books = session.exec(
        select(OrderItem.book_title, func.sum(OrderItem.quantity))
        .group_by(OrderItem.book_title)
        .order_by(func.sum(OrderItem.quantity).desc())
        .limit(5)
    ).all()

    for title, sold in books:
        ws3.append([title, sold])

    # =========================
    # Sheet 4 — Top Customers
    # =========================
    ws4 = wb.create_sheet("Top Customers")
    ws4.append(["Email", "Orders", "Total Spent"])

    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    customers = session.exec(
        select(User.email, func.count(Order.id), func.sum(Order.total))
        .join(Order, Order.user_id == User.id)
        .where(Order.status == "paid")
        .group_by(User.email)
        .order_by(func.sum(Order.total).desc())
        .limit(5)
    ).all()

    for email, orders, spent in customers:
        ws4.append([email, orders, spent])

    # =========================
    # Sheet 5 — Category Sales
    # =========================
    ws5 = wb.create_sheet("Category Sales")
    ws5.append(["Category", "Units Sold"])

    for cell in ws5[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    categories = session.exec(
        select(Category.name, func.sum(OrderItem.quantity))
        .join(Book, Book.category_id == Category.id)
        .join(OrderItem, OrderItem.book_id == Book.id)
        .group_by(Category.name)
    ).all()

    for name, sold in categories:
        ws5.append([name, sold])

    # =========================
    # Save file
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"analytics_{datetime.utcnow().date()}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
